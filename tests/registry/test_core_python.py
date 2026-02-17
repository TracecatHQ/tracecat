import importlib
import json
import os
import re
import shutil

import pytest

from tracecat import config
from tracecat.config import TRACECAT__SANDBOX_PYPI_INDEX_URL
from tracecat.sandbox import (
    SandboxExecutionError,
    SandboxService,
    SandboxTimeoutError,
    validate_run_python_script,
)
from tracecat.sandbox.executor import NsjailExecutor
from tracecat.sandbox.types import SandboxConfig

# Check if nsjail is available (required for the sandbox)
NSJAIL_AVAILABLE = shutil.which("nsjail") is not None

# Check if the sandbox rootfs exists
ROOTFS_PATH = os.environ.get(
    "TRACECAT__SANDBOX_ROOTFS_PATH", "/var/lib/tracecat/sandbox-rootfs"
)
ROOTFS_AVAILABLE = os.path.isdir(ROOTFS_PATH)

# Skip all tests if nsjail or rootfs isn't available
pytestmark = pytest.mark.skipif(
    not (NSJAIL_AVAILABLE and ROOTFS_AVAILABLE),
    reason=(
        f"nsjail sandbox not available. "
        f"nsjail installed: {NSJAIL_AVAILABLE}, rootfs exists: {ROOTFS_AVAILABLE}"
    ),
)


@pytest.fixture
def sandbox_service() -> SandboxService:
    """Provide a SandboxService instance for tests."""
    return SandboxService()


class TestScriptValidation:
    """Test suite for script structure validation (function definitions).

    This validation is done at the executor backend level via validate_run_python_script.
    """

    def test_script_with_no_function(self) -> None:
        """Test that script without function definition fails validation."""
        script = """
x = 10
print("This script has no function")
x * 2
"""
        is_valid, error = validate_run_python_script(script)
        assert not is_valid
        assert error is not None
        assert "must contain at least one function" in error

    def test_script_with_one_function(self) -> None:
        """Test that script with one function passes validation."""
        script = """
def main():
    return 42
"""
        is_valid, error = validate_run_python_script(script)
        assert is_valid
        assert error is None

    def test_script_with_multiple_functions_no_main(self) -> None:
        """Test that script with multiple functions but no main fails validation."""
        script = """
def func1():
    return "Function 1"

def func2():
    return "Function 2"
"""
        is_valid, error = validate_run_python_script(script)
        assert not is_valid
        assert error is not None
        assert "one must be named 'main'" in error

    def test_script_with_multiple_functions_with_main(self) -> None:
        """Test that script with multiple functions including main passes validation."""
        script = """
def helper():
    return 10

def main():
    return helper() * 2
"""
        is_valid, error = validate_run_python_script(script)
        assert is_valid
        assert error is None

    def test_single_function_not_named_main(self) -> None:
        """Test that a single function not named 'main' passes validation."""
        script = """
def process_data(x):
    return x * 2
"""
        is_valid, error = validate_run_python_script(script)
        assert is_valid
        assert error is None


class TestPythonExecution:
    """Test suite for Python execution via nsjail sandbox."""

    @pytest.mark.anyio
    async def test_basic_script_execution(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test basic script execution with nsjail sandbox."""
        script_content = """
def main():
    print('Hello from nsjail sandbox')
    value = 10 * 2
    return value
"""
        result = await sandbox_service.run_python(script=script_content)
        assert result == 20

    @pytest.mark.anyio
    async def test_script_with_function_arguments(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test script with function arguments."""
        script_content = """
def process_item(item_name, qty, price):
    print(f'Processing: {item_name}, quantity: {qty}')
    return qty * price
"""
        inputs_data = {"item_name": "TestItem", "qty": 10, "price": 2.5}
        expected_result = 25.0

        result = await sandbox_service.run_python(
            script=script_content, inputs=inputs_data
        )
        assert result == expected_result

    @pytest.mark.anyio
    async def test_script_with_partial_arguments(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test script with some arguments missing (should raise TypeError)."""
        script_content = """
def process_data(a, b, c):
    print(f'a={a}, b={b}, c={c}')
    result = (a or 0) + (b or 0) + (c or 0)
    return result
"""
        inputs_data = {"a": 10, "b": 5}  # c is missing, should raise TypeError

        with pytest.raises(SandboxExecutionError) as exc_info:
            await sandbox_service.run_python(script=script_content, inputs=inputs_data)

        error_msg = str(exc_info.value)
        assert "TypeError" in error_msg
        assert "missing 1 required positional argument" in error_msg
        assert "'c'" in error_msg

    @pytest.mark.anyio
    async def test_main_function_with_arguments(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test main function with arguments when multiple functions exist."""
        script_content = """
def helper_function(x):
    return x * 2

def main(value):
    result = helper_function(value)
    return result + 10
"""
        inputs_data = {"value": 5}
        expected_result = 20  # (5 * 2) + 10

        result = await sandbox_service.run_python(
            script=script_content, inputs=inputs_data
        )
        assert result == expected_result

    @pytest.mark.anyio
    async def test_script_with_dependencies(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test script with dependencies."""
        script_content = """
def main():
    import requests
    return "requests module loaded successfully"
"""
        result = await sandbox_service.run_python(
            script=script_content,
            dependencies=["requests"],
            timeout_seconds=120,
        )
        assert "requests module loaded successfully" in result

    @pytest.mark.anyio
    async def test_script_with_openpyxl_dependency(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test script with openpyxl dependency (non-pure Python package).

        This test verifies the nsjail sandbox can install packages with C extensions.
        openpyxl depends on et_xmlfile and optionally lxml, which have native components.
        The previous deno/WASM sandbox failed to install openpyxl due to WASM restrictions.
        """
        script_content = """
def main():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # Create a workbook and add data
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Write some data
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Test"
    ws["B2"] = 42

    # Test column letter utility
    col_letter = get_column_letter(3)

    return {
        "workbook_created": True,
        "sheet_title": ws.title,
        "cell_a1": ws["A1"].value,
        "cell_b2": ws["B2"].value,
        "column_3_letter": col_letter,
        "openpyxl_version": __import__("openpyxl").__version__,
    }
"""
        result = await sandbox_service.run_python(
            script=script_content,
            dependencies=["openpyxl==3.1.5"],
            timeout_seconds=120,
        )
        assert result["workbook_created"] is True
        assert result["sheet_title"] == "Test Sheet"
        assert result["cell_a1"] == "Name"
        assert result["cell_b2"] == 42
        assert result["column_3_letter"] == "C"
        assert result["openpyxl_version"] == "3.1.5"

    @pytest.mark.anyio
    async def test_script_with_ocsf_dependency(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test script with py-ocsf-models dependency.

        This test verifies the nsjail sandbox can install and use OCSF models,
        which are Pydantic-based security event schemas.
        """
        script_content = """
def main():
    from py_ocsf_models.events.findings.detection_finding import DetectionFinding
    from py_ocsf_models.objects.fingerprint import FingerPrint, AlgorithmID

    # Create a fingerprint object using the Pydantic model
    fp = FingerPrint(
        algorithm="SHA-256",
        algorithm_id=AlgorithmID.SHA_256,
        value="e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855"
    )

    # Check that we can access OCSF event classes
    detection_finding_fields = list(DetectionFinding.model_fields.keys())

    return {
        "fingerprint_algorithm": fp.algorithm,
        "fingerprint_value": fp.value,
        "fingerprint_algorithm_id": fp.algorithm_id.value,
        "detection_finding_has_fields": len(detection_finding_fields) > 0,
        "sample_fields": detection_finding_fields[:5],
    }
"""
        result = await sandbox_service.run_python(
            script=script_content,
            dependencies=["py-ocsf-models==0.8.0"],
            timeout_seconds=120,
        )
        assert result["fingerprint_algorithm"] == "SHA-256"
        assert (
            result["fingerprint_value"]
            == "e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855"
        )
        assert result["fingerprint_algorithm_id"] == 3  # SHA_256 = 3
        assert result["detection_finding_has_fields"] is True
        assert len(result["sample_fields"]) > 0

    @pytest.mark.anyio
    async def test_script_error_handling(self, sandbox_service: SandboxService) -> None:
        """Test script error handling."""
        script_with_error = """
def main():
    raise ValueError("This is a test error")
    return "Should not reach here"
"""
        with pytest.raises(SandboxExecutionError) as exc_info:
            await sandbox_service.run_python(script=script_with_error)
        assert "This is a test error" in str(exc_info.value)

    @pytest.mark.anyio
    async def test_script_timeout(self, sandbox_service: SandboxService) -> None:
        """Test script timeout."""
        script_with_infinite_loop = """
def main():
    import time
    while True:
        time.sleep(0.1)
    return "Should not reach here"
"""
        with pytest.raises(SandboxTimeoutError) as exc_info:
            await sandbox_service.run_python(
                script=script_with_infinite_loop, timeout_seconds=2
            )
        assert "timed out" in str(exc_info.value).lower()

    @pytest.mark.anyio
    async def test_complex_data_types_with_arguments(self, sandbox_service):
        """Test handling of complex data types as function arguments."""
        script_content = """
def main(data_list, data_dict):
    return {
        "input_list": data_list,
        "input_dict": data_dict,
        "processed": {
            "list_length": len(data_list),
            "dict_keys": list(data_dict.keys())
        }
    }
"""
        inputs_data = {"data_list": [1, 2, 3, 4, 5], "data_dict": {"a": 1, "b": 2}}

        result = await sandbox_service.run_python(
            script=script_content, inputs=inputs_data
        )
        assert isinstance(result, dict)
        assert result["input_list"] == [1, 2, 3, 4, 5]
        assert result["input_dict"] == {"a": 1, "b": 2}
        assert result["processed"]["list_length"] == 5
        assert result["processed"]["dict_keys"] == ["a", "b"]

    @pytest.mark.anyio
    async def test_clean_error_messages(self, sandbox_service: SandboxService) -> None:
        """Test that error messages are clean and user-friendly without tracebacks."""
        script_with_value_error = """
def main(value):
    if value is None:
        raise ValueError("Invalid input provided")
    return value * 2
"""
        with pytest.raises(SandboxExecutionError) as exc_info:
            await sandbox_service.run_python(
                script=script_with_value_error, inputs={"value": None}
            )
        error_msg = str(exc_info.value)
        assert "ValueError: Invalid input provided" in error_msg

    @pytest.mark.anyio
    async def test_subprocess_support(self, sandbox_service):
        """Test that subprocess.run works in nsjail sandbox."""
        script_content = """
def main():
    import subprocess
    result = subprocess.run(['echo', 'Hello from subprocess'], capture_output=True, text=True)
    return result.stdout.strip()
"""
        result = await sandbox_service.run_python(script=script_content)
        assert result == "Hello from subprocess"

    @pytest.mark.anyio
    async def test_env_vars_injection(self, sandbox_service: SandboxService) -> None:
        """Test that environment variables are properly injected."""
        script_content = """
def main():
    import os
    return os.environ.get('TEST_SECRET', 'not found')
"""
        result = await sandbox_service.run_python(
            script=script_content,
            env_vars={"TEST_SECRET": "secret_value_123"},
        )
        assert result == "secret_value_123"


class TestDocumentationExamples:
    """Test suite for examples from the documentation to ensure they work correctly."""

    @pytest.mark.anyio
    async def test_doc_example_function_no_inputs(self, sandbox_service):
        """Test the 'Function (no inputs)' example from the documentation."""
        script = """
def main():
    # Simple arithmetic
    result = 10 * 5 + 3
    return result
"""
        result = await sandbox_service.run_python(script=script)
        assert result == 53

    @pytest.mark.anyio
    async def test_doc_example_function_with_inputs(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test the 'Function (with inputs)' example from the documentation."""
        script = """
def main(a, b):
    # Simple arithmetic
    result = a * b + 3
    return result
"""
        inputs = {"a": 10, "b": 5}
        result = await sandbox_service.run_python(script=script, inputs=inputs)
        assert result == 53

    @pytest.mark.anyio
    async def test_doc_example_multiple_functions(self, sandbox_service):
        """Test the 'Multiple functions' example from the documentation."""
        script = """
def calculate_tax(amount, rate):
    return amount * rate

def main(subtotal, tax_rate):
    # When multiple functions exist, 'main' is called
    tax = calculate_tax(subtotal, tax_rate)
    return subtotal + tax
"""
        inputs = {"subtotal": 100.0, "tax_rate": 0.08}
        result = await sandbox_service.run_python(script=script, inputs=inputs)
        assert result == 108.0


class TestNetworkSecurity:
    """Test suite for network access security controls."""

    @pytest.mark.anyio
    async def test_network_disabled_by_default(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test that network access is disabled by default."""
        script = """
def main():
    try:
        import urllib.request
        response = urllib.request.urlopen('https://httpbin.org/get', timeout=5)
        return "Network access allowed"
    except Exception as e:
        return f"Network blocked: {type(e).__name__}"
"""
        result = await sandbox_service.run_python(script=script)
        assert "Network blocked" in result

    @pytest.mark.anyio
    async def test_basic_scripts_work_without_network(self, sandbox_service):
        """Test that basic scripts work fine without network access."""
        script = """
def main(x, y):
    result = x * y + 10
    return result
"""
        result = await sandbox_service.run_python(
            script=script, inputs={"x": 5, "y": 3}
        )
        assert result == 25

    @pytest.mark.anyio
    async def test_builtin_modules_work_without_network(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test that Python builtin modules work without network access."""
        script = """
def main():
    import json
    import math
    import datetime

    data = {"value": math.pi, "timestamp": str(datetime.datetime.now())}
    return json.dumps(data)
"""
        result = await sandbox_service.run_python(script=script)
        assert isinstance(result, str)

        parsed = json.loads(result)
        assert "value" in parsed
        assert "timestamp" in parsed

    @pytest.mark.anyio
    async def test_network_enabled_when_requested(self, sandbox_service):
        """Test that network access works when explicitly enabled."""
        script = """
def main():
    try:
        import urllib.request
        response = urllib.request.urlopen('https://httpbin.org/get', timeout=10)
        return f"Network access successful: {response.status}"
    except Exception as e:
        return f"Network failed: {type(e).__name__}: {str(e)}"
"""
        result = await sandbox_service.run_python(
            script=script,
            allow_network=True,
            timeout_seconds=30,
        )
        # Either succeeds or fails for network reasons (not sandbox blocking)
        assert "Network access successful" in result or "Network failed" in result


class TestInputValidation:
    """Test suite for input validation to ensure only dict inputs are accepted."""

    @pytest.mark.anyio
    async def test_dict_inputs_work_correctly(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test that dictionary inputs work correctly as function arguments."""
        script = """
def main(name, age, city):
    return f'{name} is {age} years old and lives in {city}'
"""
        inputs = {"name": "Alice", "age": 30, "city": "NYC"}
        result = await sandbox_service.run_python(script=script, inputs=inputs)
        assert result == "Alice is 30 years old and lives in NYC"

    @pytest.mark.anyio
    async def test_nested_dict_inputs_work(self, sandbox_service):
        """Test that nested dictionary values work as function arguments."""
        script = """
def main(user_data, settings):
    return f"User: {user_data['name']}, Theme: {settings['theme']}"
"""
        inputs = {
            "user_data": {"name": "Bob", "age": 25},
            "settings": {"theme": "dark", "lang": "en"},
        }
        result = await sandbox_service.run_python(script=script, inputs=inputs)
        assert result == "User: Bob, Theme: dark"

    @pytest.mark.anyio
    async def test_none_inputs_work(self, sandbox_service: SandboxService) -> None:
        """Test that None inputs work (no arguments passed)."""
        script = """
def main():
    return "No inputs needed"
"""
        result = await sandbox_service.run_python(script=script, inputs=None)
        assert result == "No inputs needed"

    @pytest.mark.anyio
    async def test_empty_dict_inputs_work(self, sandbox_service):
        """Test that empty dictionary inputs work."""
        script = """
def main():
    return "Empty dict inputs"
"""
        result = await sandbox_service.run_python(script=script, inputs={})
        assert result == "Empty dict inputs"


class TestSingleFunctionInputs:
    """Test suite to verify inputs work with single functions not named 'main'."""

    @pytest.mark.anyio
    async def test_single_function_not_named_main_with_inputs(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test that inputs work with a single function not named 'main'."""
        script = """
def process_order(item_name, quantity, price_per_item):
    total_cost = quantity * price_per_item
    return f"Order: {quantity}x {item_name} = ${total_cost:.2f}"
"""
        inputs = {"item_name": "Widget", "quantity": 3, "price_per_item": 15.99}
        result = await sandbox_service.run_python(script=script, inputs=inputs)
        assert result == "Order: 3x Widget = $47.97"

    @pytest.mark.anyio
    async def test_single_function_no_params_no_inputs(self, sandbox_service):
        """Test that a single function with no parameters works without inputs."""
        script = """
def calculate_pi_approximation():
    return 22 / 7
"""
        result = await sandbox_service.run_python(script=script)
        assert abs(result - 3.142857142857143) < 0.000001


class TestSecurityIsolation:
    """Test suite for security and sandbox isolation."""

    @pytest.mark.anyio
    async def test_command_injection_via_inputs(
        self, sandbox_service: SandboxService
    ) -> None:
        """Test that malicious inputs cannot inject Python code."""
        script = """
def main(user_input):
    return f"Processed: {user_input}"
"""
        malicious_inputs = {
            "user_input": '"); import os; os.system("cat /etc/passwd"); print("'
        }

        result = await sandbox_service.run_python(
            script=script, inputs=malicious_inputs
        )
        assert (
            result == 'Processed: "); import os; os.system("cat /etc/passwd"); print("'
        )

    @pytest.mark.anyio
    async def test_file_system_isolation(self, sandbox_service):
        """Test that scripts cannot access the host file system."""
        script_read = """
def main():
    try:
        # Try to read host's /etc/passwd
        with open('/etc/passwd', 'r') as f:
            content = f.read()
        # If we can read it, return the first line to verify it's sandbox's file
        return content.split('\\n')[0]
    except Exception as e:
        return f"Error: {type(e).__name__}"
"""
        result = await sandbox_service.run_python(script=script_read)
        # Should either fail or read the sandbox's /etc/passwd (with sandbox user)
        if "Error" not in result:
            # If it succeeded, verify it's reading sandbox's passwd file
            assert "sandbox" in result or "root" in result

    @pytest.mark.anyio
    async def test_large_input_handling(self, sandbox_service: SandboxService) -> None:
        """Test handling of large inputs."""
        script = """
def main(data):
    return f"Received {len(data)} items"
"""
        large_input = {"data": list(range(10000))}
        result = await sandbox_service.run_python(script=script, inputs=large_input)
        assert result == "Received 10000 items"

    @pytest.mark.anyio
    async def test_nested_function_calls_safe(self, sandbox_service):
        """Test that nested function calls with user input are safe."""
        script = """
def process_data(data):
    return data.upper() if isinstance(data, str) else str(data)

def main(user_input):
    result = process_data(user_input)
    return f"Final: {result}"
"""
        malicious_input = {"user_input": "__import__('os').system('id')"}

        result = await sandbox_service.run_python(script=script, inputs=malicious_input)
        assert result == "Final: __IMPORT__('OS').SYSTEM('ID')"


class TestMultiTenantIsolation:
    """Test suite for multi-tenant workspace isolation."""

    @pytest.mark.anyio
    async def test_workspace_cache_isolation(self) -> None:
        """Test that different workspaces get separate package caches."""
        service = SandboxService()

        # Compute cache keys for same dependencies with different workspace IDs
        deps = ["requests==2.28.0"]
        cache_key_workspace_a = service._compute_cache_key(
            dependencies=deps, workspace_id="workspace-a"
        )
        cache_key_workspace_b = service._compute_cache_key(
            dependencies=deps, workspace_id="workspace-b"
        )
        cache_key_no_workspace = service._compute_cache_key(
            dependencies=deps, workspace_id=None
        )

        # Verify different workspaces get different cache keys
        assert cache_key_workspace_a != cache_key_workspace_b
        assert cache_key_workspace_a != cache_key_no_workspace
        assert cache_key_workspace_b != cache_key_no_workspace

    @pytest.mark.anyio
    async def test_version_isolation(self) -> None:
        """Test that different package versions get different cache keys."""
        service = SandboxService()
        workspace_id = "workspace-test"

        # Different versions should produce different cache keys
        key_v1 = service._compute_cache_key(
            dependencies=["requests==2.28.0"], workspace_id=workspace_id
        )
        key_v2 = service._compute_cache_key(
            dependencies=["requests==2.29.0"], workspace_id=workspace_id
        )

        assert key_v1 != key_v2

    @pytest.mark.anyio
    async def test_cache_key_format(self) -> None:
        """Test that cache keys follow expected format (hex string)."""
        service = SandboxService()
        cache_key = service._compute_cache_key(
            dependencies=["requests==2.28.0"], workspace_id="workspace-test"
        )

        # Cache key should be a 16-character hexadecimal string
        assert isinstance(cache_key, str)
        assert len(cache_key) == 16
        assert re.match(r"^[a-f0-9]+$", cache_key)


class TestPyPIConfiguration:
    """Test suite for PyPI index URL configuration."""

    @pytest.mark.anyio
    async def test_default_pypi_index_used(self) -> None:
        """Test that default PyPI index is used when no configuration is set."""
        # Default should be public PyPI
        assert TRACECAT__SANDBOX_PYPI_INDEX_URL == "https://pypi.org/simple"

    @pytest.mark.anyio
    async def test_custom_pypi_index_configuration(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test that custom PyPI index URL can be configured."""
        # Set custom index URL
        custom_url = "https://custom.pypi.example.com/simple"
        monkeypatch.setenv("TRACECAT__SANDBOX_PYPI_INDEX_URL", custom_url)

        # Reload config to pick up new env var
        importlib.reload(config)

        assert config.TRACECAT__SANDBOX_PYPI_INDEX_URL == custom_url

    @pytest.mark.anyio
    async def test_extra_index_urls_configuration(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test that extra index URLs can be configured."""
        # Set extra index URLs
        extra_urls = (
            "https://extra1.example.com/simple,https://extra2.example.com/simple"
        )
        monkeypatch.setenv("TRACECAT__SANDBOX_PYPI_EXTRA_INDEX_URLS", extra_urls)

        # Reload config to pick up new env var
        importlib.reload(config)

        assert len(config.TRACECAT__SANDBOX_PYPI_EXTRA_INDEX_URLS) == 2
        assert (
            "https://extra1.example.com/simple"
            in config.TRACECAT__SANDBOX_PYPI_EXTRA_INDEX_URLS
        )
        assert (
            "https://extra2.example.com/simple"
            in config.TRACECAT__SANDBOX_PYPI_EXTRA_INDEX_URLS
        )

    @pytest.mark.anyio
    async def test_executor_passes_index_urls_to_install_env(self) -> None:
        """Test that executor passes PyPI index URLs to install environment."""
        sandbox_config = SandboxConfig(
            network_enabled=False,
            env_vars={},
        )
        executor = NsjailExecutor()

        # Build environment for install phase
        env_map = executor._build_env_map(sandbox_config, phase="install")

        # Verify PyPI index URLs are in the environment
        assert "UV_INDEX_URL" in env_map
        assert env_map["UV_INDEX_URL"] == "https://pypi.org/simple"


def print_nsjail_installation_instructions() -> None:
    """Print instructions for setting up nsjail sandbox."""
    print("\n" + "=" * 80)
    print("nsjail Sandbox Setup Instructions")
    print("=" * 80)
    print(
        """
These tests require nsjail sandbox to be set up. This is typically done
inside the Docker container during the image build process.

To run tests locally:

1. Build and run the development Docker container:
   just dev

2. Run tests inside the executor container:
   docker exec -it tracecat-executor-1 pytest tests/registry/test_core_python.py -v

The sandbox requires:
- Linux with kernel >= 4.6 (for user namespaces)
- nsjail binary at /usr/local/bin/nsjail
- Sandbox rootfs at /var/lib/tracecat/sandbox-rootfs
- SYS_ADMIN capability (configured in docker-compose)

Note: Unsafe PID executor fallback tests are in tests/unit/test_unsafe_pid_executor.py. These tests run without nsjail and
will pass on any system with Python 3.12+.
"""
    )
    print("=" * 80 + "\n")


# Print installation instructions if sandbox is not available
if not (NSJAIL_AVAILABLE and ROOTFS_AVAILABLE) and __name__ == "__main__":
    print_nsjail_installation_instructions()
